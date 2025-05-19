
import streamlit as st
import pandas as pd
import fitz  # PyMuPDF
import re
from io import BytesIO
from openpyxl import load_workbook

st.title("Preencher Planilha com Dados do PDF de Julgamento")
st.markdown("Envie o PDF do julgamento e a planilha modelo para preenchimento automático.")

# Upload dos arquivos
pdf_file = st.file_uploader("PDF do julgamento:", type="pdf")
xlsx_file = st.file_uploader("Planilha modelo (.xlsx):", type="xlsx")

# Regex para extrair dados
padrao_geral = re.compile(
    r"Item (\d+)[^\n]*?\n.*?"
    r"Aceito e Habilitado.*?para\s+(.*?),\s+CNPJ\s+(\d{2}\.\d{3}\.\d{3}/\d{4}-\d{2}),.*?"
    r"melhor lance:\s*R\$ ([\d\.,]+).*?/ R\$ ([\d\.,]+)",
    re.DOTALL
)

def extrair_dados_pdf(file):
    dados = {}
    with fitz.open(stream=file.read(), filetype="pdf") as doc:
        for page in doc:
            texto = page.get_text()
            for match in padrao_geral.findall(texto):
                item = int(match[0])
                empresa = match[1].replace("\n", " ").strip()
                valor_unit = float(match[3].replace(".", "").replace(",", "."))
                valor_total = float(match[4].replace(".", "").replace(",", "."))
                dados[item] = {
                    "fornecedor": empresa,
                    "valor_unitario": valor_unit,
                    "valor_total": valor_total
                }
    return dados

if pdf_file and xlsx_file:
    st.info("🔍 Extraindo dados do PDF e preenchendo planilha...")

    dados_pdf = extrair_dados_pdf(pdf_file)
    wb = load_workbook(xlsx_file)
    ws = wb.active

    for row in ws.iter_rows(min_row=4, max_col=9):
        try:
            item_val = row[0].value
            item_num = int(str(item_val).replace(".", "").strip())
            if item_num in dados_pdf:
                row[6].value = dados_pdf[item_num]["valor_unitario"]
                row[7].value = dados_pdf[item_num]["valor_total"]
                row[8].value = dados_pdf[item_num]["fornecedor"]
            else:
                row[8].value = "Fracassado e/ou Deserto"
        except:
            row[8].value = "Fracassado e/ou Deserto"

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    st.success("✅ Planilha preenchida com sucesso!")
    st.download_button("🔗 Baixar planilha preenchida", output, file_name="Planilha_modelo_preenchida_FINAL.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
